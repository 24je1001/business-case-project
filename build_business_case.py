#!/usr/bin/env python3
"""
Automated Business Case & Financial Model Builder
---------------------------------------------------
An executive-level CLI tool that generates high-end corporate financial workbooks
complete with dynamic macro metrics, advanced formatting, and smoothed data visualizations.
"""

import argparse
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Executive Style Constants (Premium Corporate Palette)
# ---------------------------------------------------------------------------
FONT_FAMILY = "Segoe UI"

# Typography
TITLE_FONT = Font(name=FONT_FAMILY, size=16, bold=True, color="1B365D")
HEADER_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
SUBHEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="1B365D")
LABEL_FONT = Font(name=FONT_FAMILY, size=10, bold=False, color="333333")
BLUE_INPUT = Font(name=FONT_FAMILY, size=10, bold=True, color="1D4ED8")      # Professional Royal Blue for inputs
BLACK_FORMULA = Font(name=FONT_FAMILY, size=10, color="000000")              # Clean black for calculated outputs
KPI_LABEL_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="FFFFFF")
KPI_VALUE_FONT = Font(name=FONT_FAMILY, size=14, bold=True, color="1B365D")
NOTE_FONT = Font(name=FONT_FAMILY, size=9, italic=True, color="6B7280")

# Color Fills
HEADER_FILL = PatternFill("solid", start_color="1B365D")        # Midnight Navy
SUBHEADER_FILL = PatternFill("solid", start_color="E8EEF5")     # Soft Ice Blue
KPI_FILL = PatternFill("solid", start_color="2C3E50")           # Deep Slate
INPUT_FILL = PatternFill("solid", start_color="FAF9F5")         # Premium Soft Alabaster Cream

# Borders & Alignments
THIN_BORDER = Side(style="thin", color="D1D5DB")                 # Clean muted gray lines
BOX = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# Number Formats
CURRENCY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'
YEAR_FMT = '0'


def parse_args():
    p = argparse.ArgumentParser(description="Generate an executive-ready business case Excel asset.")
    p.add_argument("--project-name", default="Enterprise Cloud & AI Transformation",
                    help="Name of the investment project")
    p.add_argument("--client-name", default="Global Enterprise Corp",
                    help="Target client identifier")
    p.add_argument("--implementation-cost", type=float, default=450000,
                    help="Core platform rollout execution fees")
    p.add_argument("--iot-sensor-cost", type=float, default=125000,
                    help="Edge infrastructure & equipment procurement hardware")
    p.add_argument("--annual-savings", type=float, default=260000,
                    help="Gross operational yield efficiency gains (Year 1)")
    p.add_argument("--savings-growth", type=float, default=0.04,
                    help="Compounding annual yield scale optimization factor")
    p.add_argument("--annual-maintenance", type=float, default=35000,
                    help="Sustaining software licensing & support operational overhead")
    p.add_argument("--discount-rate", type=float, default=0.085,
                    help="Corporate hurdle/wacc discount rate")
    p.add_argument("--years", type=int, default=5,
                    help="Length of the forecast structure")
    p.add_argument("--output", default="business_case.xlsx",
                    help="Target destination filename")
    return p.parse_args()


def style_row(ws, row, start_col, end_col, fill, font, alignment=CENTER, height=24):
    ws.row_dimensions[row].height = height
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        if fill: cell.fill = fill
        if font: cell.font = font
        if alignment: cell.alignment = alignment
        cell.border = BOX


def build_workbook(args):
    wb = Workbook()

    # =========================================================
    # TAB 1: Strategic Assumptions
    # =========================================================
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a.sheet_view.showGridLines = False
    ws_a.column_dimensions["A"].width = 38
    ws_a.column_dimensions["B"].width = 22
    ws_a.column_dimensions["C"].width = 50

    ws_a["A1"] = args.project_name
    ws_a["A1"].font = TITLE_FONT
    ws_a["A2"] = f"Financial Strategy Proposal | Prepared for: {args.client_name}"
    ws_a["A2"].font = LABEL_FONT

    ws_a["A4"] = "Model Configuration & Architecture Parameters"
    style_row(ws_a, 4, 1, 3, HEADER_FILL, HEADER_FONT, LEFT, height=26)
    ws_a.merge_cells("A4:C4")

    headers = ["Financial Driver Vector", "Base Value Assumption", "Strategic Context & Accounting Treatment"]
    for i, h in enumerate(headers, start=1):
        cell = ws_a.cell(row=5, column=i, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.border = BOX
        cell.alignment = CENTER
    ws_a.row_dimensions[5].height = 22

    rows = [
        ("Platform Implementation Capital ($)", args.implementation_cost, "Capitalized system design, architecture, and deployment services."),
        ("Edge Infrastructure Hardware ($)", args.iot_sensor_cost, "Procurement of sensor arrays, endpoints, and physical network nodes."),
        ("Initial Operational Savings ($)", args.annual_savings, "Targeted Year 1 gross utility, process efficiency, and labor optimization."),
        ("Savings Scaling Optimization Rate", args.savings_growth, "Compounding annual efficiency coefficient from automated process maturity."),
        ("Sustaining Maintenance Overhead ($)", args.annual_maintenance, "Recurring service legal agreements, upgrades, and support run-rate."),
        ("Corporate Discount Rate (WACC)", args.discount_rate, "Corporate hurdle rate optimized for present-value cash stream factoring."),
        ("Strategic Valuation Horizon (Years)", args.years, "Active duration modeled for infrastructure run lifecycle analysis."),
    ]

    for i, (label, value, note) in enumerate(rows):
        r = 6 + i
        ws_a.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws_a.cell(row=r, column=1).border = BOX
        ws_a.cell(row=r, column=1).alignment = LEFT
        
        vcell = ws_a.cell(row=r, column=2, value=value)
        vcell.font = BLUE_INPUT
        vcell.fill = INPUT_FILL
        vcell.border = BOX
        vcell.alignment = CENTER
        
        if "Rate" in label or "WACC" in label:
            vcell.number_format = PCT_FMT
        elif "Horizon" in label:
            vcell.number_format = YEAR_FMT
        else:
            vcell.number_format = CURRENCY_FMT
            
        ncell = ws_a.cell(row=r, column=3, value=note)
        ncell.font = NOTE_FONT
        ncell.border = BOX
        ncell.alignment = LEFT
        ws_a.row_dimensions[r].height = 22

    IMPL_COST = "Assumptions!$B$6"
    IOT_COST = "Assumptions!$B$7"
    SAVINGS_Y1 = "Assumptions!$B$8"
    SAVINGS_GROWTH = "Assumptions!$B$9"
    MAINTENANCE = "Assumptions!$B$10"
    DISCOUNT_RATE = "Assumptions!$B$11"

    ws_a["A15"] = "Standard Convention Notice: Royal Blue metrics denote manual assumptions input variables. Dark black outputs signify automated formula paths."
    ws_a["A15"].font = NOTE_FONT
    ws_a.merge_cells("A15:C15")

    # =========================================================
    # TAB 2: Valuation Model
    # =========================================================
    ws = wb.create_sheet("Financial Model")
    ws.sheet_view.showGridLines = False

    n_years = args.years
    last_col = 1 + n_years + 1  
    ws.column_dimensions["A"].width = 34
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    ws["A1"] = "Projected Valuation Ledger & Cash Flow Analysis"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")

    # Year header structure
    year_row = 3
    ws.cell(row=year_row, column=1, value="Timeline Index").font = SUBHEADER_FONT
    ws.cell(row=year_row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=year_row, column=1).border = BOX
    ws.cell(row=year_row, column=1).alignment = LEFT

    for y in range(0, n_years + 1):
        col = 2 + y
        c = ws.cell(row=year_row, column=col, value=f"Year {y}" if y > 0 else "Year 0 (CapEx)")
        c.font = SUBHEADER_FONT
        c.alignment = CENTER
        c.fill = SUBHEADER_FILL
        c.border = BOX
    ws.row_dimensions[year_row].height = 24

    # Row 4: Capex Outlays
    r = 4
    ws.cell(row=r, column=1, value="Initial Capital Investment").font = LABEL_FONT
    for y in range(0, n_years + 1):
        col = 2 + y
        cell = ws.cell(row=r, column=col)
        cell.value = f"=-({IMPL_COST}+{IOT_COST})" if y == 0 else 0
        cell.font = BLACK_FORMULA
        cell.number_format = CURRENCY_FMT
        cell.border = BOX
    style_row(ws, r, 1, last_col, fill=None, font=None, alignment=None, height=22)
    INVESTMENT_ROW = r

    # Row 5: Efficiency Yield
    r = 5
    ws.cell(row=r, column=1, value="Gross Operational Savings").font = LABEL_FONT
    for y in range(0, n_years + 1):
        col = 2 + y
        cell = ws.cell(row=r, column=col)
        if y == 0: cell.value = 0
        elif y == 1: cell.value = f"={SAVINGS_Y1}"
        else: cell.value = f"={get_column_letter(col-1)}{r}*(1+{SAVINGS_GROWTH})"
        cell.font = BLACK_FORMULA
        cell.number_format = CURRENCY_FMT
        cell.border = BOX
    style_row(ws, r, 1, last_col, fill=None, font=None, alignment=None, height=22)
    SAVINGS_ROW = r

    # Row 6: OPEX Maintenance
    r = 6
    ws.cell(row=r, column=1, value="Sustaining Support Costs").font = LABEL_FONT
    for y in range(0, n_years + 1):
        col = 2 + y
        cell = ws.cell(row=r, column=col)
        cell.value = 0 if y == 0 else f"=-{MAINTENANCE}"
        cell.font = BLACK_FORMULA
        cell.number_format = CURRENCY_FMT
        cell.border = BOX
    style_row(ws, r, 1, last_col, fill=None, font=None, alignment=None, height=22)
    MAINT_ROW = r

    # Row 7: Net Cash Flows
    r = 7
    ws.cell(row=r, column=1, value="Net Annual Cash Flow").font = SUBHEADER_FONT
    for y in range(0, n_years + 1):
        col = 2 + y
        col_l = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={col_l}{INVESTMENT_ROW}+{col_l}{SAVINGS_ROW}+{col_l}{MAINT_ROW}"
        cell.font = SUBHEADER_FONT
        cell.number_format = CURRENCY_FMT
        cell.fill = SUBHEADER_FILL
        cell.border = BOX
    ws.row_dimensions[r].height = 24
    NET_CF_ROW = r

    # Row 8: Cumulative Nominal Position
    r = 8
    ws.cell(row=r, column=1, value="Cumulative Nominal Cash Flow").font = LABEL_FONT
    for y in range(0, n_years + 1):
        col = 2 + y
        col_l = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        if y == 0: cell.value = f"={col_l}{NET_CF_ROW}"
        else: cell.value = f"={get_column_letter(col-1)}{r}+{col_l}{NET_CF_ROW}"
        cell.font = BLACK_FORMULA
        cell.number_format = CURRENCY_FMT
        cell.border = BOX
    style_row(ws, r, 1, last_col, fill=None, font=None, alignment=None, height=22)
    CUM_CF_ROW = r

    # Row 9: Discount Vector Factor
    r = 9
    ws.cell(row=r, column=1, value="Present Value Factor").font = LABEL_FONT
    for y in range(0, n_years + 1):
        col = 2 + y
        col_l = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"=1/(1+{DISCOUNT_RATE})^{y}"
        cell.font = BLACK_FORMULA
        cell.number_format = '0.0000'
        cell.border = BOX
    style_row(ws, r, 1, last_col, fill=None, font=None, alignment=None, height=22)
    DISC_ROW = r

    # Row 10: Present Value Flow
    r = 10
    ws.cell(row=r, column=1, value="Discounted Net Cash Flow (PV)").font = LABEL_FONT
    for y in range(0, n_years + 1):
        col = 2 + y
        col_l = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        cell.value = f"={col_l}{NET_CF_ROW}*{col_l}{DISC_ROW}"
        cell.font = BLACK_FORMULA
        cell.number_format = CURRENCY_FMT
        cell.border = BOX
    style_row(ws, r, 1, last_col, fill=None, font=None, alignment=None, height=22)
    PV_ROW = r

    # Row 11: Cumulative Present Value
    r = 11
    ws.cell(row=r, column=1, value="Cumulative Present Value (NPV Trend)").font = LABEL_FONT
    for y in range(0, n_years + 1):
        col = 2 + y
        col_l = get_column_letter(col)
        cell = ws.cell(row=r, column=col)
        if y == 0: cell.value = f"={col_l}{PV_ROW}"
        else: cell.value = f"={get_column_letter(col-1)}{r}+{col_l}{PV_ROW}"
        cell.font = BLACK_FORMULA
        cell.number_format = CURRENCY_FMT
        cell.border = BOX
    style_row(ws, r, 1, last_col, fill=None, font=None, alignment=None, height=22)
    CUM_PV_ROW = r

    # Reference cell calculation coordinates
    first_yr1_col = get_column_letter(3)          
    last_yr_col = get_column_letter(2 + n_years)   
    year0_col = get_column_letter(2)

    # =========================================================
    # Executive Scorecard Blocks (Dashboard)
    # =========================================================
    kpi_row = r + 3
    ws.cell(row=kpi_row, column=1, value="Executive Investment Scorecard Metrics").font = HEADER_FONT
    style_row(ws, kpi_row, 1, min(5, last_col), KPI_FILL, HEADER_FONT, LEFT, height=26)
    ws.merge_cells(start_row=kpi_row, start_column=1, end_row=kpi_row, end_column=min(5, last_col))

    label_row = kpi_row + 1
    value_row = kpi_row + 2

    kpis = [
        ("Net Present Value (NPV)", f"=NPV({DISCOUNT_RATE},{first_yr1_col}{NET_CF_ROW}:{last_yr_col}{NET_CF_ROW})+{year0_col}{NET_CF_ROW}", CURRENCY_FMT),
        ("Internal Return Multiplier (ROI)", f"=SUM({first_yr1_col}{NET_CF_ROW}:{last_yr_col}{NET_CF_ROW})/-{year0_col}{NET_CF_ROW}", PCT_FMT),
        ("Calculated Payback Lifecycle", 
         (f'=IFERROR(MATCH(TRUE,INDEX({year0_col}{CUM_CF_ROW}:{last_yr_col}{CUM_CF_ROW}>0,0))-2'
          f'+(-INDEX({year0_col}{CUM_CF_ROW}:{last_yr_col}{CUM_CF_ROW},MATCH(TRUE,INDEX({year0_col}{CUM_CF_ROW}:{last_yr_col}{CUM_CF_ROW}>0,0))-1))'
          f'/INDEX({year0_col}{NET_CF_ROW}:{last_yr_col}{NET_CF_ROW},MATCH(TRUE,INDEX({year0_col}{CUM_CF_ROW}:{last_yr_col}{CUM_CF_ROW}>0,0))),"Out of Range")'), 
         '0.00" Yrs"'),
        ("Total Net Capital Delta", f"=SUM({first_yr1_col}{SAVINGS_ROW}:{last_yr_col}{SAVINGS_ROW})+SUM({first_yr1_col}{MAINT_ROW}:{last_yr_col}{MAINT_ROW})", CURRENCY_FMT),
    ]

    for i, (label, formula, fmt) in enumerate(kpis):
        col = 1 + i
        lcell = ws.cell(row=label_row, column=col, value=label)
        lcell.font = KPI_LABEL_FONT
        lcell.fill = PatternFill("solid", start_color="34495E")
        lcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        lcell.border = BOX
        ws.row_dimensions[label_row].height = 28

        vcell = ws.cell(row=value_row, column=col, value=formula)
        vcell.font = KPI_VALUE_FONT
        vcell.alignment = CENTER
        vcell.border = BOX
        vcell.number_format = fmt
        vcell.fill = PatternFill("solid", start_color="F8FAFC")
        ws.row_dimensions[value_row].height = 30

    # Conditional highlights for primary NPV cell status
    NPV_CELL = f"B{value_row}"
    ws.conditional_formatting.add(NPV_CELL, CellIsRule(operator="greaterThanOrEqual", formula=["0"], font=Font(color="15803D", bold=True)))
    ws.conditional_formatting.add(NPV_CELL, CellIsRule(operator="lessThan", formula=["0"], font=Font(color="B91C1C", bold=True)))

    # =========================================================
    # Premium Data Visualization Integration
    # =========================================================
    chart = LineChart()
    chart.title = "Investment Return Horizons & Cumulative Trajectory"
    chart.style = 13  # High-contrast premium profile theme
    chart.y_axis.title = "Valuation Streams ($)"
    chart.x_axis.title = "Valuation Increments"
    chart.height = 14
    chart.width = 24

    # Explicitly configure axes visible parameters
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.y_axis.numFmt = "$#,##0"
    chart.legend.position = "b"  # Move legend to the bottom area to leave chart space clear

    cats = Reference(ws, min_col=2, max_col=last_col, min_row=year_row, max_row=year_row)
    data_net = Reference(ws, min_col=1, max_col=last_col, min_row=NET_CF_ROW, max_row=NET_CF_ROW)
    data_cum = Reference(ws, min_col=1, max_col=last_col, min_row=CUM_CF_ROW, max_row=CUM_CF_ROW)
    data_cumpv = Reference(ws, min_col=1, max_col=last_col, min_row=CUM_PV_ROW, max_row=CUM_PV_ROW)

    chart.add_data(data_net, titles_from_data=True, from_rows=True)
    chart.add_data(data_cum, titles_from_data=True, from_rows=True)
    chart.add_data(data_cumpv, titles_from_data=True, from_rows=True)
    chart.set_categories(cats)

    # Apply modern curve smoothing to trend chart vectors
    for series in chart.series:
        series.marker = Marker(symbol="circle", size=5)
        series.smooth = True

    anchor_row = value_row + 3
    ws.add_chart(chart, f"A{anchor_row}")

    # CRITICAL INSTRUCTION: Unfreeze all viewport windows explicitly
    ws.freeze_panes = None

    # =========================================================
    # TAB 3: Operational Notes
    # =========================================================
    ws_n = wb.create_sheet("Notes")
    ws_n.sheet_view.showGridLines = False
    ws_n.column_dimensions["A"].width = 110
    ws_n["A1"] = "Operational Guidelines & Architectural Design Criteria"
    ws_n["A1"].font = TITLE_FONT
    
    notes = [
        "1. Interconnected Architecture: Modifications applied inside the yellow 'Assumptions' tab propagate instantly down calculation layers.",
        "2. Capital Outlay Alignment: Baseline software platforms and ancillary node machinery components aggregate directly into Year 0 outflows.",
        "3. Precision Payback Formula: Calculated fields use true linear index estimation points to establish target breakeven metrics cleanly.",
        "4. Secured Document Notice: Ensure you select 'Enable Editing' in the Microsoft Excel notification ribbons to unlock full local evaluation calculations."
    ]
    for i, note in enumerate(notes):
        cell = ws_n.cell(row=3 + i, column=1, value=note)
        cell.font = Font(name=FONT_FAMILY, size=10)
        ws_n.row_dimensions[3 + i].height = 24

    return wb


def main():
    args = parse_args()
    wb = build_workbook(args)
    wb.save(args.output)
    print(f"Executive business case successfully built to file target: {args.output}")


if __name__ == "__main__":
    main()
    